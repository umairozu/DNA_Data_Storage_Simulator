import math
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

import numpy as np
from openpyxl import load_workbook


R = 8.31446261815324  # J / (mol K)
SEC_PER_WEEK = 7 * 24 * 3600
SEC_PER_YEAR = 365 * 24 * 3600


@dataclass
class Arrhenius_decay:
    """
    First-order DNA decay model calibrated from the workbook used for
    Li et al., "A compact cassette tape for DNA-based data storage"
    (Science Advances, 2025).

    Calibration procedure
    ---------------------
    1. For each temperature and time point, average the measured DNA
       concentration replicates.
    2. Fit ln(mean concentration) versus time:

           ln C(t) = -k t + ln C0

       so that k = -slope.

       D-DNA uses weeks 0, 1, 2.
       E-DNA uses weeks 0, 1, 2, 3.

    3. Fit ln(k) versus 1/T:

           ln k = ln A - Ea / (R T)

       giving Ea from the slope and ln(A) from the intercept.

    At the experimentally measured temperatures (60, 65, 70 C), k()
    returns the directly fitted experimental k value. At other
    temperatures, k() uses the Arrhenius fit for interpolation or
    extrapolation.
    """

    k_dDNA: Dict[int, float]  # temperature C -> measured decay constant, s^-1
    k_eDNA: Dict[int, float]
    ea_d: float               # J/mol
    ea_e: float               # J/mol
    lnA_dDNA: float
    lnA_eDNA: float

    # Row blocks in RawData.xlsx, sheet "Arrhenius Calculate".
    blocks = {
        0: {60: (2, 7), 65: (2, 7), 70: (2, 7)},
        1: {60: (8, 13), 65: (14, 19), 70: (20, 25)},
        2: {60: (26, 31), 65: (32, 37), 70: (38, 43)},
        3: {60: (44, 49), 65: (50, 55), 70: (56, 61)},
    }

    @staticmethod
    def numeric_cells(ws, col: str, r1: int, r2: int) -> List[float]:
        values: List[float] = []
        for row in range(r1, r2 + 1):
            value = ws[f"{col}{row}"].value
            if isinstance(value, (int, float)):
                value = float(value)
                if math.isfinite(value):
                    values.append(value)
        return values

    @classmethod
    def mean_concentration(
        cls,
        ws,
        col: str,
        week: int,
        temp_C: int,
    ) -> Optional[float]:
        r1, r2 = cls.blocks[week][temp_C]
        values = cls.numeric_cells(ws, col, r1, r2)
        if not values:
            return None
        return float(np.mean(values))

    @staticmethod
    def fit_k(ln_means: np.ndarray, time_s: np.ndarray) -> float:
        """
        Fit ln(mean concentration) = -k*t + intercept and return k in s^-1.
        """
        if len(ln_means) != len(time_s) or len(ln_means) < 2:
            raise ValueError("At least two matched time/concentration points are required")

        slope, _intercept = np.polyfit(time_s, ln_means, 1)
        k = float(-slope)
        if not math.isfinite(k) or k <= 0.0:
            raise ValueError(f"Invalid fitted decay constant k={k}")
        return k

    @classmethod
    def fit_k_from_means(
        cls,
        ws,
        col: str,
        temp_C: int,
        max_week: int,
    ) -> float:

        times: List[float] = []
        ln_means: List[float] = []

        for week in range(0, max_week + 1):
            mean_value = cls.mean_concentration(ws, col, week, temp_C)
            if mean_value is None:
                continue
            if mean_value <= 0.0:
                raise ValueError(
                    f"Non-positive mean concentration for col={col}, "
                    f"temp={temp_C} C, week={week}: {mean_value}"
                )

            times.append(float(week * SEC_PER_WEEK))
            ln_means.append(math.log(mean_value))

        if len(times) < 2:
            raise ValueError(
                f"Insufficient valid data to fit k for col={col}, temp={temp_C} C"
            )

        return cls.fit_k(
            np.asarray(ln_means, dtype=float),
            np.asarray(times, dtype=float),
        )

    @staticmethod
    def fit_arrhenius(k_table: Dict[int, float]) -> Tuple[float, float]:
        """
        Fit ln(k) = ln(A) - Ea/(R*T).

        Returns (Ea, lnA)
        """
        if len(k_table) < 2:
            raise ValueError("At least two temperatures are required for Arrhenius fitting")

        temperatures_K = []
        ln_k = []

        for temp_C in sorted(k_table):
            k_value = float(k_table[temp_C])
            if k_value <= 0.0 or not math.isfinite(k_value):
                raise ValueError(f"Invalid k at {temp_C} C: {k_value}")

            T = 273.15 + float(temp_C)
            temperatures_K.append(T)
            ln_k.append(math.log(k_value))

        x = 1.0 / np.asarray(temperatures_K, dtype=float)
        y = np.asarray(ln_k, dtype=float)

        slope, intercept = np.polyfit(x, y, 1)
        ea = float(-slope * R)
        lnA = float(intercept)

        if ea <= 0.0 or not math.isfinite(ea) or not math.isfinite(lnA):
            raise ValueError(f"Invalid Arrhenius fit: Ea={ea}, lnA={lnA}")

        return ea, lnA

    @classmethod
    def from_xlsx(cls, xlsx_path: str) -> "Arrhenius_decay":
        """Load RawData.xlsx and build the calibrated D-DNA/E-DNA model."""
        wb = load_workbook(xlsx_path, data_only=True, read_only=True)
        try:
            ws = wb["Arrhenius Calculate"]

            temperatures = (60, 65, 70)

            # D-DNA: weeks 0, 1, 2 only.
            k_dDNA = {
                temp: cls.fit_k_from_means(
                    ws,
                    col="K",
                    temp_C=temp,
                    max_week=2,
                )
                for temp in temperatures
            }

            # E-DNA: weeks 0, 1, 2, 3.
            k_eDNA = {
                temp: cls.fit_k_from_means(
                    ws,
                    col="Q",
                    temp_C=temp,
                    max_week=3,
                )
                for temp in temperatures
            }
        finally:
            wb.close()

        ea_d, lnA_dDNA = cls.fit_arrhenius(k_dDNA)
        ea_e, lnA_eDNA = cls.fit_arrhenius(k_eDNA)

        return cls(
            k_dDNA=k_dDNA,
            k_eDNA=k_eDNA,
            ea_d=ea_d,
            ea_e=ea_e,
            lnA_dDNA=lnA_dDNA,
            lnA_eDNA=lnA_eDNA,
        )

    def k(self, temp_C: float, encapsulated: bool) -> float:
        """
        Return the first-order decay constant k.

        Exact experimental temperatures (60/65/70 C) use the directly fitted
        workbook value. Other temperatures use the Arrhenius fit.
        """
        temp_C = float(temp_C)
        T = 273.15 + temp_C
        if T <= 0.0:
            raise ValueError("Temperature must be above absolute zero")

        table = self.k_eDNA if bool(encapsulated) else self.k_dDNA

        rounded_temp = int(round(temp_C))
        if (
            rounded_temp in table
            and math.isclose(temp_C, float(rounded_temp), rel_tol=0.0, abs_tol=1e-12)
        ):
            return float(table[rounded_temp])

        if bool(encapsulated):
            ea = self.ea_e
            lnA = self.lnA_eDNA
        else:
            ea = self.ea_d
            lnA = self.lnA_dDNA

        k_value = math.exp(lnA - ea / (R * T))
        if not math.isfinite(k_value) or k_value < 0.0:
            raise ValueError(
                f"Invalid Arrhenius decay constant at {temp_C} C: {k_value}"
            )
        return k_value

    def remaining_dna_frac(
        self,
        temp_C: float,
        encapsulated: bool,
        week: float,
    ) -> float:
        """
        Remaining bulk DNA fraction after storage:

            C(t) / C0 = exp(-k*t)
        """
        week = float(week)
        if week < 0.0:
            raise ValueError("Storage duration in weeks must be >= 0")
        if week == 0.0:
            return 1.0

        t = SEC_PER_WEEK * week
        exponent = -self.k(temp_C, encapsulated) * t

        # math.exp safely underflows to 0 for sufficiently negative values,
        # which is appropriate for effectively complete molecular loss.
        fraction = math.exp(exponent)
        return float(min(max(fraction, 0.0), 1.0))

    def lost_dna_frac(
        self,
        temp_C: float,
        encapsulated: bool,
        week: float,
    ) -> float:
        """Fraction of bulk DNA lost during storage."""
        return 1.0 - self.remaining_dna_frac(temp_C, encapsulated, week)

    def half_life(self, temp_C: float, encapsulated: bool) -> float:
        """Concentration half-life at the requested temperature, in years."""
        k_value = self.k(temp_C, encapsulated)
        if k_value <= 0.0:
            return math.inf
        return (math.log(2.0) / k_value) / SEC_PER_YEAR

    def protection_ratio(self, temp_C: float) -> float:
        """
        E-DNA / D-DNA bulk decay-rate ratio at a temperature.

        Values < 1 mean the encapsulated material decays more slowly.
        """
        k_d = self.k(temp_C, encapsulated=False)
        k_e = self.k(temp_C, encapsulated=True)
        if k_d <= 0.0:
            raise ZeroDivisionError("D-DNA decay constant is zero")
        return k_e / k_d


    @classmethod
    def fit_k_linregress(
        cls,
        ws,
        col: str,
        temp: int,
        max_week: int,
    ) -> float:
        return cls.fit_k_from_means(ws, col, temp, max_week)

    @classmethod
    def fit_k_lingress(
        cls,
        ws,
        col: str,
        temp: int,
        max_week: int,
    ) -> float:
        return cls.fit_k_from_means(ws, col, temp, max_week)


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description="Inspect the calibrated DNA-tape Arrhenius decay model"
    )
    parser.add_argument("xlsx", nargs="?", default="RawData.xlsx")
    parser.add_argument("--temp", type=float, default=20.0)
    args = parser.parse_args()

    model = Arrhenius_decay.from_xlsx(args.xlsx)

    print("Measured decay constants (s^-1)")
    print("D-DNA:", model.k_dDNA)
    print("E-DNA:", model.k_eDNA)
    print()
    print(f"D-DNA Ea: {model.ea_d / 1000.0:.3f} kJ/mol")
    print(f"E-DNA Ea: {model.ea_e / 1000.0:.3f} kJ/mol")
    print(f"D-DNA ln(A): {model.lnA_dDNA:.6f}")
    print(f"E-DNA ln(A): {model.lnA_eDNA:.6f}")
    print()
    print(f"At {args.temp:g} C:")
    print(f"  D-DNA k: {model.k(args.temp, False):.8e} s^-1")
    print(f"  E-DNA k: {model.k(args.temp, True):.8e} s^-1")
    print(f"  D-DNA half-life: {model.half_life(args.temp, False):.6g} years")
    print(f"  E-DNA half-life: {model.half_life(args.temp, True):.6g} years")
    print(f"  E/D bulk decay ratio: {model.protection_ratio(args.temp):.8g}")